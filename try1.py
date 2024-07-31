import cv2
import argparse
import csv
import json
import numpy as np
from tqdm import tqdm
from os.path import exists
import os
import random
import datetime
import logging
from segment_anything import sam_model_registry as sam_registry
from mobile_sam import sam_model_registry as mobile_sam_registry
from countanything.countanything import SamAutomaticMaskGenerator
import matplotlib.pyplot as plt
import time
import wandb
import torch
from collections import defaultdict
from sklearn.cluster import DBSCAN
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from tqdm import tqdm

# Function to show annotations
def show_anns(anns, image, output_path):
    if len(anns) == 0:
        return
    sorted_anns = sorted(anns, key=(lambda x: x['area']), reverse=True)
    fig, ax = plt.subplots(figsize=(10, 10))
    ax.imshow(image)
    ax.set_autoscale_on(False)
    for ann in sorted_anns:
        x0, y0, w, h = ann['bbox']
        ax.add_patch(plt.Rectangle((x0, y0), w, h, edgecolor='green', facecolor=(0,0,0,0), lw=2))
        ax.scatter([x0 + w // 2], [y0 + h // 2], color='green', marker='*', s=10, edgecolor='white', linewidth=1.25)
    ax.axis('off')
    plt.savefig(output_path, bbox_inches='tight', pad_inches=0)
    plt.close(fig)

def draw_and_save_rectangles(image_path, bounding_boxes, output_dir):

    image = cv2.imread(image_path)
    
    # Create a copy of the image to draw on
    output_image = image.copy()
    
    # Draw rectangles on the image
    for box in bounding_boxes:
        if len(box) != 4:
            raise ValueError("Each bounding box must be a list of 4 elements [x, y, w, h].")
        
        x, y, w, h = map(int, box)  # Ensure coordinates are integers
        cv2.rectangle(output_image, (x, y), (x + w, y + h), (0, 255, 0), 2)  # Draw rectangle
    
    # Save the modified image with rectangles drawn on it
    filename = os.path.basename(image_path)
    output_path = os.path.join(output_dir, filename)
    
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    #cv2.imwrite(output_path, output_image)
    print(f"Image with rectangles saved to {output_path}")


def draw_and_save_circles_and_boxes(image_path, circles, boxes, output_dir):
    # Read the image
    image = cv2.imread(image_path)
    
    # Create a copy of the image to draw on
    output_image = image.copy()
    
    # Draw circles on the image
    for circle in circles:
        center_x, center_y, radius = circle
        cv2.circle(output_image, (center_x, center_y), radius, (0, 255, 0), 2)  # Draw circle
    
    # Draw bounding boxes on the image
    for box in boxes:
        x, y, width, height = box
        top_left = (x, y)
        bottom_right = (x + width, y + height)
        cv2.rectangle(output_image, top_left, bottom_right, (255, 0, 0), 2)  # Draw bounding box
    
    # Save the modified image with circles and boxes drawn on it
    filename = os.path.basename(image_path)
    output_path = os.path.join(output_dir, filename)
    
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    cv2.imwrite(output_path, output_image)
    print(f"Image with circles and boxes saved to {output_path}")

def show_prompt(anns):
    if len(anns) == 0:
        return
    sorted_anns = sorted(anns, key=(lambda x: x['area']), reverse=True)
    ax = plt.gca()
    ax.set_autoscale_on(False)
    for ann in sorted_anns:
        x0, y0, w, h = ann['bbox']
        ax.add_patch(plt.Rectangle((x0, y0), w, h, edgecolor='green', facecolor=(0,0,0,0), lw=2))
        ax.scatter([x0+w//2], [y0+h//2], color='green', marker='*', s=10, edgecolor='white', linewidth=1.25)

def calculate_iou(bbox1, bbox2):
    """
    Calculate the Intersection over Union (IoU) between two bounding boxes.
    
    Arguments:
    bbox1: List representing the coordinates of the first bounding box in the format [x, y, width, height].
    bbox2: List representing the coordinates of the second bounding box in the format [x, y, width, height].
    
    Returns:
    iou: Intersection over Union (IoU) value between the two bounding boxes.
    """
    # Convert bbox1 to [x1, y1, x2, y2] format
    x1_bbox1, y1_bbox1, x2_bbox1, y2_bbox1 = bbox1[0], bbox1[1], bbox1[0] + bbox1[2], bbox1[1] + bbox1[3]
    # Convert bbox2 to [x1, y1, x2, y2] format
    x1_bbox2, y1_bbox2, x2_bbox2, y2_bbox2 = bbox2[0], bbox2[1], bbox2[0] + bbox2[2], bbox2[1] + bbox2[3]
    
    # Calculate the coordinates of the intersection rectangle
    x1_intersection = max(x1_bbox1, x1_bbox2)
    y1_intersection = max(y1_bbox1, y1_bbox2)
    x2_intersection = min(x2_bbox1, x2_bbox2)
    y2_intersection = min(y2_bbox1, y2_bbox2)
    
    # Calculate the area of intersection rectangle
    intersection_area = max(0, x2_intersection - x1_intersection + 1) * max(0, y2_intersection - y1_intersection + 1)
    
    # Calculate the area of each bounding box
    area_b1 = (x2_bbox1 - x1_bbox1 + 1) * (y2_bbox1 - y1_bbox1 + 1)
    area_b2 = (x2_bbox2 - x1_bbox2 + 1) * (y2_bbox2 - y1_bbox2 + 1)
    
    # Calculate the Union area
    union_area = area_b1 + area_b2 - intersection_area
    
    # Calculate IoU
    iou = intersection_area / union_area if union_area > 0 else 0
    return iou



def detect_and_categorize_circles(img_path, epsilon, min_samples, prompt_number):
    print(f"Running detect_and_categorize_circles with epsilon: {epsilon}, min_samples: {min_samples}, and {prompt_number} prompts")
    print(img_path)
    # Read the image
    image = cv2.imread(img_path)
    
    # Convert the image to grayscale
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    # Apply Gaussian Blur to reduce noise
    blurred = cv2.GaussianBlur(gray, (9, 9), 2)
    
    # Apply Canny edge detection
    edges = cv2.Canny(blurred, 50, 150)

    # Detect circles in the image
    dp = 1
    min_dist = 20
    param1 = 50
    param2 = 0.7
    min_radius = 0
    max_radius = 0
    circles = cv2.HoughCircles(
        gray, 
        cv2.HOUGH_GRADIENT_ALT, 
        dp=dp, 
        minDist=min_dist, 
        param1=param1, 
        param2=param2, 
        minRadius=min_radius, 
        maxRadius=max_radius
    )

    #print(f"Detected circles: {circles}")
    # Check if any circles are detected
    if circles is None:
        return [], []

    circles = np.uint16(np.around(circles))
    
    # Extract radii from circles
    radii = circles[0, :, 2].reshape(-1, 1)
    
    # Apply DBSCAN to cluster radii
    dbscan = DBSCAN(eps=epsilon, min_samples=min_samples)
    labels = dbscan.fit_predict(radii)
    
    # Categorize circles based on DBSCAN labels
    categorized_circles = defaultdict(list)
    for label, circle in zip(labels, circles[0, :]):
        if label != -1:  # Ignore noise points
            categorized_circles[label].append(circle)
    
    # Find the most populous category
    if not categorized_circles:
        return [], []
    
    most_populous_category = max(categorized_circles, key=lambda k: len(categorized_circles[k]))
    most_common_circles = categorized_circles[most_populous_category]
    
    # Create bounding boxes for the closest circles to the center
    bboxes = []
    selected_circles = []
    num_circles = min(prompt_number, len(most_common_circles))
    
    def do_bounding_boxes_overlap(bbox1, bbox2):
        x1, y1, w1, h1 = bbox1
        x2, y2, w2, h2 = bbox2
        return not (x1 + w1 < x2 or x2 + w2 < x1 or y1 + h1 < y2 or y2 + h2 < y1)
    
    for i in range(num_circles):
        center_x, center_y, radius = most_common_circles[i]
        x = center_x - radius
        y = center_y - radius
        width = 2 * radius
        height = 2 * radius
        bbox = [x, y, width, height]  # Bounding box in [x, y, width, height] format

        # Check for overlap with existing bounding boxes
        if not any(do_bounding_boxes_overlap(bbox, existing_bbox) for existing_bbox in bboxes):
            bboxes.append(bbox)
            selected_circles.append(most_common_circles[i])
    
    print(bboxes)
    return bboxes, selected_circles


# Combine both registries
combined_registry = {**mobile_sam_registry, **sam_registry}

# Argument parser setup
parser = argparse.ArgumentParser(description="Few Shot Counting Evaluation code")
parser.add_argument("-dp",  "--data_path", type=str, default='counting_pipe.v1i.coco/test/images', help="Path to the FSC147 dataset")
parser.add_argument("-mt",  "--model_type", type=str, default="vit_l", help="model type")
parser.add_argument("-mp",  "--model_path", type=str, default="Models/sam_vit_h_4b8939.pth", help="path to trained model")
parser.add_argument("-op",  "--output_path", type=str, default="Benchmark_Result_default_path", help="path to Output")
parser.add_argument("-io",  "--test_iou_threshold", type=float, default=0.75, help="Set testing Iou Thresh'hold")
parser.add_argument("-ip",  "--pred_iou_threshold", type=float, default=0.8, help="Set AutoMaskGen prediction Iou Thresh'hold")
parser.add_argument("-st",  "--sim_thresh", type=float, default=0.8, help="Set Similarity Thresh'hold")
parser.add_argument("-np",  "--prompt_num", type=int, default=3, help="chose number of prompts")
parser.add_argument("-p1",  "--param1", type=int, default=75, help="Higher threshold for the Canny edge detector")
parser.add_argument("-p2",  "--param2", type=float, default=0.85, help="confidence")
parser.add_argument("-eps",  "--epsilon", type=int, default=20, help="epsilon")
parser.add_argument("-ms",  "--min_samples", type=int, default=3, help="min_samples")
parser.add_argument("-v",   "--viz", type=bool, default=True, help="whether to visualize")
parser.add_argument("-d",   "--device", default='0', help='assign device')

args = parser.parse_args()

# Assign arguments to variables
pred_iou_thresh = args.pred_iou_threshold
sim_thresh = args.sim_thresh
data_path = args.data_path
anno_file = os.path.join(data_path, 'merged_annotations.json')
im_dir = os.path.join(data_path)
test_iou_threshold = args.test_iou_threshold
benchmark_dir = args.output_path

os.makedirs(benchmark_dir, exist_ok=True)

prompt_type="hough_circle"
model_name=args.model_type
num_files = len(os.listdir(im_dir))
prompt_number=args.prompt_num

param1 = args.param1
param2 = args.param2
epsilon=args.epsilon
min_samples= args.min_samples
debug = True

os.environ['CUDA_VISIBLE_DEVICES'] = args.device.strip()
device = 'cuda'
sam = combined_registry[args.model_type](checkpoint=args.model_path)
sam.to(device=device)

# Initialize wandb
wandb.init(
    project="countany benchmark houghcircle",
    config={
        "Iou_thresh": pred_iou_thresh,
        "sim_thresh": sim_thresh,
        "Number of images": num_files,
        "Model": model_name,
        "prompting": prompt_type,
        "number of prompts": prompt_number,
        "confidence": param2
    }
)

mask_generator = SamAutomaticMaskGenerator(model=sam, min_mask_region_area=25, sim_thresh=sim_thresh, pred_iou_thresh=pred_iou_thresh)

min_radius = 10
max_radius = 300
min_dist = 30
dp = 0.7

data = []

# Loop through all subdirectories in benchmark_dir
for sub_dir in tqdm([d for d in os.listdir(data_path) if os.path.isdir(os.path.join(data_path, d))]):
    sub_dir_path = os.path.join(data_path, sub_dir)
    print(f"Processing subdirectory: {sub_dir_path}")
    
    img_file = None
    txt_file = None
    
    # Find the image file and txt file in the subdirectory
    for file in os.listdir(sub_dir_path):
        if file.lower().endswith(('.png', '.jpg', '.jpeg')):
            if file.lower() == 'result_image_with_bboxes.jpg':
                continue  # Skip this file
            img_file = file
        elif file.lower().endswith('.txt'):
            txt_file = file
    
    if not img_file or not txt_file:
        print(f"Missing image or text file in {sub_dir_path}")
        continue
    
    img_path = os.path.join(sub_dir_path, img_file)
    txt_path = os.path.join(sub_dir_path, txt_file)
    
    with open(txt_path, 'r') as f:
        gt_count = int(f.read().strip())
    
    print(f"Image file: {img_file}, Ground truth count: {gt_count}")
    
    try:
        bounding_boxes, circles_to_draw = detect_and_categorize_circles(img_path, epsilon, min_samples, prompt_number)
        #print(f"Bounding boxes detected: {bounding_boxes}")
    except Exception as e:
        print(f"An error occurred while processing {img_path}: {e}")
        continue
    
    if len(bounding_boxes) == 0:
        print("bounding_boxes is empty")
        continue
    
    input_boxes = [[bbox[0], bbox[1], bbox[0] + bbox[2], bbox[1] + bbox[3]] for bbox in bounding_boxes]
    input_boxes = np.array(input_boxes, dtype=np.float64)

    print(f"Number of input boxes: {len(input_boxes)}")

    image = cv2.imread(img_path)
    if image is None:
        print(f"Failed to read image: {img_path}")
        continue

    masks = mask_generator.generate(image, input_boxes)
    
    pred_cnt = len(masks)
    print(f"Predicted count: {pred_cnt}")

    # Draw bounding boxes on the image
    for mask in masks:
        print("mask:", mask['bbox'])
        x, y, w, h = map(int, mask['bbox'])
        cv2.rectangle(image, (x, y), (x + w, y + h), (0, 255, 0), 2)  # Draw rectangle

    # Save the image with bounding boxes
    
    result_img_path = os.path.join(sub_dir_path, "result_image_with_bboxes.jpg")
    cv2.imwrite(result_img_path, image)
    print(f"Image with bounding boxes saved to: {result_img_path}")

    # Append data for Excel
    data.append({
        "Subdirectory": sub_dir,
        "GT Count": gt_count,
        "Pred Count": pred_cnt,
        "Original Image": img_path,
        "Result Image": result_img_path
    })
timestamp=datetime.datetime.now()
# Save data to Excel with images
excel_file_path = os.path.join(benchmark_dir, f"benchmark_results{timestamp}.xlsx")
wb = Workbook()
ws = wb.active

# Add headers
ws.append(["Subdirectory", "GT Count", "Pred Count", "Original Image", "Result Image"])

for row, item in enumerate(data, start=2):
    ws[f"A{row}"] = item["Subdirectory"]
    ws[f"B{row}"] = item["GT Count"]
    ws[f"C{row}"] = item["Pred Count"]
    
    # Insert original image
    if os.path.exists(item["Original Image"]):
        img = ExcelImage(item["Original Image"])
        img.anchor = f"D{row}"
        ws.add_image(img)
    else:
        print(f"Original image not found: {item['Original Image']}")
    
    # Insert result image
    result_img_path = item["Result Image"]
    if os.path.exists(result_img_path):
        result_img = ExcelImage(result_img_path)
        result_img.anchor = f"E{row}"
        ws.add_image(result_img)
    else:
        print(f"Result image not found: {result_img_path}")

wb.save(excel_file_path)
print(f"Excel saved to {excel_file_path}")